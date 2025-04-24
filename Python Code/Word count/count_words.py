import xml.etree.ElementTree as ET
from collections import Counter
import re
import csv
from openpyxl import load_workbook

def extract_text_from_element(element):
    text_content = []
    if element.text:
        text_content.append(element.text)
    for child in element:
        text_content.extend(extract_text_from_element(child))
    if element.tail:
        text_content.append(element.tail)
    return text_content

def count_words_in_xml(file_path):
    # Parse the XML file
    tree = ET.parse(file_path)
    root = tree.getroot()

    # Extract text content from all elements
    text_content = extract_text_from_element(root)

    # Join all text content into a single string
    text_string = ' '.join(text_content)

    # Use regex to find all words (alphanumeric sequences)
    words = re.findall(r'\b\w+\b', text_string.lower())

    # Count occurrences of each word
    word_counts = Counter(words)

    return word_counts

def count_words_in_csv(file_path):
    text_content = []
    with open(file_path, newline='', encoding='utf-8', errors='ignore') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            text_content.extend(row)

    # Join all text content into a single string
    text_string = ' '.join(text_content)

    # Use regex to find all words (alphanumeric sequences)
    words = re.findall(r'\b\w+\b', text_string.lower())

    # Count occurrences of each word
    word_counts = Counter(words)

    return word_counts

def count_words_in_xlsx(file_path):
    text_content = []
    workbook = load_workbook(filename=file_path)
    for sheet in workbook:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    text_content.append(str(cell))

    # Join all text content into a single string
    text_string = ' '.join(text_content)

    # Use regex to find all words (alphanumeric sequences)
    words = re.findall(r'\b\w+\b', text_string.lower())

    # Count occurrences of each word
    word_counts = Counter(words)

    return word_counts

def save_word_counts_to_file(word_counts, output_file):
    # Sort word counts in reverse order of count
    sorted_word_counts = sorted(word_counts.items(), key=lambda item: item[1], reverse=True)
    
    with open(output_file, 'w',  encoding='utf-8', errors='replace') as f:
        for word, count in sorted_word_counts:
            f.write(f'{word}: {count}\n')

# Example usage for XML
# xml_file_path = 'sdn.xml'
# xml_output_file = 'sdn_xml_word_counts.txt'
# xml_word_counts = count_words_in_xml(xml_file_path)
# save_word_counts_to_file(xml_word_counts, xml_output_file)
# print(f"XML word counts have been saved to {xml_output_file} in reverse order of count.")

# Example usage for CSV
# csv_file_path = 'consolidated.csv'
# csv_output_file = 'consolidated_csv_word_counts.txt'
# csv_word_counts = count_words_in_csv(csv_file_path)
# save_word_counts_to_file(csv_word_counts, csv_output_file)
# print(f"CSV word counts have been saved to {csv_output_file} in reverse order of count.")

# Example usage for XLSX
xlsx_file_path = 'fincen.xlsx'
xlsx_output_file = 'fincen_xlsx_word_counts.txt'
xlsx_word_counts = count_words_in_xlsx(xlsx_file_path)
save_word_counts_to_file(xlsx_word_counts, xlsx_output_file)
print(f"XLSX word counts have been saved to {xlsx_output_file} in reverse order of count.")